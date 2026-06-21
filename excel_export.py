import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from extractor import InvoiceData


def generate_excel(data: InvoiceData) -> bytes:
    """
    Convert an InvoiceData object to an Excel file (bytes).
    Format and design are fixed as per requirement — do not change.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ── Styles (do not change) ────────────────────────────────────────────────
    header_fill = PatternFill(start_color="9BBB59", end_color="9BBB59", fill_type="solid")
    border_style = Side(border_style="thin", color="000000")
    border = Border(
        left=border_style, right=border_style,
        top=border_style, bottom=border_style
    )
    center_aligned = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aligned   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    right_aligned  = Alignment(horizontal="right",  vertical="center")

    # ── Detect bill type ──────────────────────────────────────────────────────
    has_igst = any(
        item.igst_amount is not None
        for item in data.line_items
    )

    # ── Headers ───────────────────────────────────────────────────────────────
    if has_igst:
        headers = [
            "S.NO", "Vendor Name", "GSTN", "Invoice Number", "Invoice Date",
            "HSN", "Product", "Qty", "Rate", "Taxable Amount",
            "IGST%", "IGST", "Total"
        ]
    else:
        headers = [
            "S.NO", "Vendor Name", "GSTN", "Invoice Number", "Invoice Date",
            "HSN", "Product", "Qty", "Rate", "Taxable Amount",
            "CGST%", "CGST", "SGST%", "SGST", "Total"
        ]

    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = header_fill
        cell.border    = border
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # ── Data rows ─────────────────────────────────────────────────────────────
    line_items = data.line_items
    num_items  = len(line_items)
    start_row  = 2
    end_row    = start_row + num_items - 1

    for i, item in enumerate(line_items):
        row_idx = start_row + i

        # First 5 columns: filled only on first row, then merged below
        if i == 0:
            ws.cell(row=row_idx, column=1, value=1)
            ws.cell(row=row_idx, column=2, value=data.seller_name or "")
            ws.cell(row=row_idx, column=3, value=data.seller_gst or "")
            ws.cell(row=row_idx, column=4, value=data.invoice_number or "")
            ws.cell(row=row_idx, column=5, value=data.invoice_date or "")

        # Line-item columns
        ws.cell(row=row_idx, column=6,  value=item.hsn_sac or "")
        ws.cell(row=row_idx, column=7,  value=item.description or "")
        ws.cell(row=row_idx, column=8,  value=item.quantity)
        ws.cell(row=row_idx, column=9,  value=item.rate)
        ws.cell(row=row_idx, column=10, value=item.taxable_amount)

        if has_igst:
            ws.cell(row=row_idx, column=11, value=item.igst_percent)
            ws.cell(row=row_idx, column=12, value=item.igst_amount)
            ws.cell(row=row_idx, column=13, value=item.total)
            last_col = 13
        else:
            ws.cell(row=row_idx, column=11, value=item.cgst_percent)
            ws.cell(row=row_idx, column=12, value=item.cgst_amount)
            ws.cell(row=row_idx, column=13, value=item.sgst_percent)
            ws.cell(row=row_idx, column=14, value=item.sgst_amount)
            ws.cell(row=row_idx, column=15, value=item.total)
            last_col = 15

        # Borders & alignment for line-item columns
        for col in range(6, last_col + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            if col == 7:
                c.alignment = left_aligned
            elif col >= 8:
                c.alignment = right_aligned
            else:
                c.alignment = center_aligned

    # ── Merge first 5 columns across all line-item rows ───────────────────────
    for col in range(1, 6):
        if num_items > 1:
            ws.merge_cells(
                start_row=start_row, start_column=col,
                end_row=end_row,     end_column=col
            )
        cell = ws.cell(row=start_row, column=col)
        cell.alignment = center_aligned
        for r in range(start_row, end_row + 1):
            ws.cell(row=r, column=col).border = border

    # ── Column widths (do not change) ─────────────────────────────────────────
    if has_igst:
        col_widths = {
            'A': 6, 'B': 35, 'C': 18, 'D': 16, 'E': 12,
            'F': 12, 'G': 50, 'H': 6, 'I': 10, 'J': 15,
            'K': 8, 'L': 10, 'M': 12
        }
    else:
        col_widths = {
            'A': 6, 'B': 35, 'C': 18, 'D': 16, 'E': 12,
            'F': 12, 'G': 50, 'H': 6, 'I': 10, 'J': 15,
            'K': 8, 'L': 10, 'M': 8, 'N': 10, 'O': 12
        }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # ── Number formatting (do not change) ─────────────────────────────────────
    amount_cols = [9, 10, 12, 14, 15] if not has_igst else [9, 10, 12, 13]
    for r in range(start_row, end_row + 1):
        for col in amount_cols:
            ws.cell(row=r, column=col).number_format = '0.00'

    # ── Return as bytes (in-memory, no file saved) ────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()